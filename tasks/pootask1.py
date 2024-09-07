from math import pi
from math import sqrt
import random
import math
import psycopg2
import json
import openpyxl
import tabula
from tabula import read_pdf
import pandas as pd
import numpy as np
from xlwt import Workbook 
from tabulate import tabulate
from django.db import connection
from datetime import datetime
import os 
import PyPDF2
from PyPDF2 import PdfReader , PdfWriter, PdfMerger
from gallery.models import *
from gallery.forms import *
from django.views.generic.edit import FormView
from .forms import FileFieldForm, FicheroForm
from .models import Fichero
from pathlib import Path
class Coffee:
    # Constructor
    def __init__(self, name, price):
        self.name = name
        self.price = float(price)
    def check_budget(self, budget):
        # Check if the budget is valid
        if not isinstance(budget, (int, float)):
            print('Enter float or int')
            exit()
            if budget < 0: 
                print('Sorry you don\'t have money') 
                exit() 
    def get_change(self, budget):
        return budget - self.price
        
    def sell(self, budget):
        self.check_budget(budget)
        if budget >= self.price:
            print(f'You can buy the {self.name} coffee')
            if budget == self.price:
                print('It\'s complete')
            else:
                print(f'Here is your change {self.get_change(budget)}$')

                exit('Thanks for your transaction')
                        
class Cookie:
	# Constructor
	def __init__(self, name, shape, chips='Chocolate'):
		# Instance attributes
		self.name = name
		self.shape = shape
		self.chips = chips

	# The object is passing itself as a parameter
	def bake(self):
		print(f'This {self.name}, is being baked with the shape {self.shape} and chips of {self.chips}')
		print('Enjoy your cookie!')

class Shape:
	def __init__(self, side1, side2):
		self.side1 = side1
		self.side2 = side2

	def get_area(self):
		return self.side1 * self.side2

	def __str__(self):
		return f'The area of this {self.__class__.__name__} is: {self.get_area()}'
 
class Rectangle(Shape): # Superclass in Parenthesis
	pass
 
class Square(Rectangle):
	def __init__(self, side):
		super().__init__(side, side)
		
class Triangle(Rectangle):
	def __init__(self, base, height):
		super().__init__(base, height)
 
	def get_area(self):
		area = super().get_area()
		return area / 2
 
class Circle(Shape):
	def __init__(self, radius):
		self.radius = radius
 
	def get_area(self):
		return pi * (self.radius ** 2)
 
class Hexagon(Rectangle):
	def __init__(self, side1):
		self.side1 = side1
	
	def get_area(self):
		return (3 * sqrt(3) * self.side1 ** 2) / 2
		
class Diccionario:

    def __init__(self, request, cmdtext):
        """Esta clase se usa para todas las operaciones con diccionarios"""
        self.request = request
        self.cmdtext = cmdtext
        
    def dict_selectuno(self):
    
        comi = "'"
        
        commandtext = f'select * from sp_selectuno ({self.cmdtext});'
        
        rawDatauno = self.exe_sp(commandtext)
        
        return rawDatauno
         
    def dict_selectodo(self,cmdtext):
    
        comi = "'"
        
        commandtext = f'select * from sp_selectodo ({self.cmdtext});'
        
        rawDatatodo = self.exe_sp(commandtext)
        
        return rawDatatodo
        
    def get_fila(self):
        
        rawData = self.dict_selectodo(self.cmdtext)
        
        thisdict = {}
        for i, fila in enumerate(rawData):
            thisdict[i] = rawData[i][0]
        
        return thisdict
        
    def rawData(self):
        
        rawData = self.dict_selectodo(self.cmdtext)
        
        return rawData
        
    def dict_ins_filas(self, num_tblu, pxcampo, pcampos, jsontextvalparam):  
        comi = "'"
        commandtext = f'select * from ins_filas({num_tblu},{comi}{pxcampo}{comi},{comi}{pcampos}{comi},{comi}{jsontextvalparam}{comi});'     
        
        return self.exe_sp(commandtext)                
    
    def dict_sp_camposinsert(self, num_tblu):  
        commandtext = f'select * from sp_camposinsert({num_tblu});'    
        
        return self.exe_sp(commandtext) 
            
    def fnt_insert(self, num_tbl, dictpart):
      eserr = False
      try:
        
        rawData = self.dict_sp_camposinsert(num_tbl) 
        xnominsert = rawData[0][0]['xcamposinsert']
        pcampos = rawData[0][0]['pcampos']
#{"camposinsert":"codigo","xcamposinsert":"xcodigo","pcampos":"codigo jsonb"} 
        constring = f'[{dictpart}]'
        constring = constring.replace("'",'"')
        
        rawData = self.dict_ins_filas(num_tbl, xnominsert, pcampos, constring) 
        #print(rawData)
        return eserr
      except Exception as e:
        print(e) 
        eserr = True  
        return eserr 
    
    def dict_sp_camposupdate(self, num_tblu):  
        commandtext = f'select * from sp_camposupdate({num_tblu});'     
        
        return self.exe_sp(commandtext) 
 
    def dict_update_filas(self, jsontextvalparam):  
        comi = "'"
        commandtext = f'select * from update_filas({jsontextvalparam});'     
        
        return self.exe_sp(commandtext) 
        
    def fnt_valdefault(self, num_tbl, rel):
        dia = datetime.today().strftime('%d-%m-%Y') 
        diastamp = datetime.now()
        valtipuser = 0
        if int(num_tbl) == 461 and int(rel) == 1:
            valtipuser = self.dict_sp_selectcount(num_tbl)[0][0] + 1
        valdefault = {1: valtipuser, 2: 'vacio', 3: dia, 4: diastamp, 7: 'vacio', 8: 0, 9: False}
        
        return valdefault[int(rel)]
            
    def builddata(self, num_tbl, cmdtext, querydict, esinsert): 
                
        rawData = self.dict_sp_camposupdate(num_tbl) 
        
        for i, filaraw in enumerate(rawData):                
            thisdict = {}         
            thisdict = rawData[i][0]
            j = 0
            for key,value in thisdict.items():
                if key == 'camposupdate':
                    fldsel = value
                elif key == 'xnomupdate':
                    xnomupdate = value
                elif key == 'xcamposinsert':
                    xnominsert = value
                elif key == 'pcampos':
                    pcampos = value
                elif key == 'pcamposout':
                    pcamposout = value
                        
            arrfldsel = fldsel.split(', ')
            
            arrfldselect = self.dict_selectuno('fldselect', cmdtext).split(", ")  
        
            arrrelacional = self.dict_selectuno('relacional', cmdtext).split(", ") 
            
            dictval = {}
            c = 0 
            for nomcamp in arrfldsel:
                rel = arrrelacional[c]
                if esinsert:
                    dictval[nomcamp] = self.fnt_valdefault(num_tbl, rel)
                else:
                    ncqdict = f'inp-{c}'
                    dictval[nomcamp] = querydict[ncqdict]
                c += 1 
             
            if len(arrfldselect) == 1 and arrfldselect[0] == 'codigo':
                data = {"codigo": dictval}
            else:
                data = dictval 
                 
            data=json.dumps(data)
            
            dataobj = {"data": data, "xnominsert": xnominsert, "xnomupdate": xnomupdate, "pcampos": pcampos, "pcamposout": pcamposout, "arrfldselect": arrfldselect, "arrrelacional": arrrelacional, "dictval": dictval}
            
            return dataobj
        
#////////////////////////////////////////////          
    def exe_sp(self, commandtext):

        print(commandtext)
    
        try:
            with connection.cursor() as cursor:
            
                cursor.execute(commandtext)
                rawData= cursor.fetchall()
               
                return rawData

        finally:
            cursor.close() 
            
class Cattbl(Diccionario):

    def __init__(self, request, cmdtext):
       super().__init__(request, cmdtext)
        
    def infoesquema(self, nomtbl):
        datainfo = self.dict_sp_information_schematblxnombre(nomtbl)
            
        return datainfo
        
    def fnt_numtype(self, typefld):
       
        numtype = {"integer": 1, "bigint": 1, "character varying": 2, "date": 3, "timestamp with time zone": 4, "jsonb": 6, "text": 7, "double precision": 8, "boolean": 9}
        
        for i,vi in numtype.items():
            if str(i) == str(typefld):
                return vi
    
    def dict_sp_information_schematblxnombre(self, nomtbl):
       #numtbl,nomtbl,tittbl
       try:    
        comi = "'"
        commandtext = f'select * from sp_information_schematblxnombre({comi}{nomtbl}{comi});' 

        rawData = super().exe_sp(commandtext) 
       
        campos = {} 
        fldselec = {}  
        fldtitulos = {}        
        relacional = {} 
        camposinsert = {}  
        camposupdate = {} 
        xnomupdate = {}  
        pcampos = {}  
        pcamposout = {}  
        xcamposinsert = {} 
        for i, fila in enumerate(rawData):
            thisdict = {}
            thisdict = rawData[i][0]
            for k,vk in thisdict.items():                   
                if k == 'column_name':
                        nomcampo = vk
                        campos[i] = vk  
                        fldselec[i] = vk   
                        fldtitulos[i] = vk  
                        camposinsert[i] = vk   
                        camposupdate[i] = vk     
                        xnomupdate[i] = f'{vk} = x.{vk}' 
                        pcamposout[i] = f'x.{vk}' 
                        xcamposinsert[i] = f'x.{vk}'                   
                elif k == 'data_type':    
                        relacional[i] = self.fnt_numtype(f'{vk}')  
                        pcampos[i] = f'{nomcampo} {vk}' 
        datacombos =  {                   
            1: campos,
            2: fldselec,
            3: fldtitulos,        
            4: relacional,
            5: camposinsert, 
            6: camposupdate,
            7: xnomupdate,
            8: pcampos,
            9: pcamposout,
           10: xcamposinsert, 
        }  
                     
        camposyy = '' 
        fldselecyy = ''  
        fldtitulosyy = ''        
        relacionalyy = '' 
        camposinsertyy = ''  
        camposupdateyy = '' 
        xnomupdateyy = ''  
        pcamposyy = ''  
        pcamposoutyy = '' 
        xcamposinsertyy = ''     
        
        for num,fldcombos in datacombos.items():
            sizeoff = len(fldcombos)
            fila = 1
            for key,value in fldcombos.items():
              if fila == sizeoff:
                if num == 1:
                    camposyy += f'{value}'
                elif num == 2:
                    fldselecyy += f'{value}'
                elif num == 3:
                    fldtitulosyy += f'{value}'
                elif num == 4:
                    relacionalyy += f'{value}'
                elif num == 5:
                    camposinsertyy += f'{value}'
                elif num == 6:
                    camposupdateyy += f'{value}'
                elif num == 7:
                    xnomupdateyy += f'{value}'
                elif num == 8:
                    pcamposyy += f'{value}'
                elif num == 9:
                    pcamposoutyy += f'{value}'
                elif num == 10:
                    xcamposinsertyy += f'{value}'
              else:
                if num == 1:
                    camposyy += f'{value}, ' 
                elif num == 2:
                    fldselecyy += f'{value}, '
                elif num == 3:
                    fldtitulosyy += f'{value}, '
                elif num == 4:
                    relacionalyy += f'{value}, '
                elif num == 5:
                    camposinsertyy += f'{value}, '
                elif num == 6:
                    camposupdateyy += f'{value}, '
                elif num == 7:
                    xnomupdateyy += f'{value}, '
                elif num == 8:
                    pcamposyy += f'{value}, '
                elif num == 9:
                    pcamposoutyy += f'{value}, '
                elif num == 10:
                    xcamposinsertyy += f'{value}, '
                    
                fila += 1
        #numtbl,nomtbl,tittbl        
        id = self.request.POST.get('numtbl')   
            
        titulo = self.request.POST.get('tittbl')   
               
        datacombos = f"INSERT INTO tareas_combos(id,nombre,titulo,campos,fldselect,fldtitulos,relacional,camposinsert,camposupdate,xnomupdate,pcampos,pcamposout,xcamposinsert) VALUES ({id},{comi}{nomtbl}{comi},{comi}{titulo}{comi},'','','','','','','','','',''); UPDATE tareas_combos SET campos='{camposyy}', fldselect='{fldselecyy}', fldtitulos='{fldtitulosyy}', relacional='{relacionalyy}', camposinsert='{camposinsertyy}', camposupdate='{camposupdateyy}', xcamposinsert='{xcamposinsertyy}', xnomupdate='{xnomupdateyy}', pcampos='{pcamposyy}', pcamposout='{pcamposoutyy}' WHERE nombre = '{nomtbl}';"
                    
        return datacombos
       except Exception as e:
         print(e) 

class Archivo(Diccionario):
    def __init__(self, request, cmdtext, extfile):
        self.extfile = extfile
        super().__init__(request, cmdtext)
         
    def filaexcel(self, fileexcel, hoja):  
        wb = openpyxl.load_workbook(fileexcel)
        # Define variable to read sheet
        ws = wb[hoja] 
        
        _row = {"titulo": [ws["A1"].value], "descri": [ws["A2"].value]}
        
        if hoja == "Hoja4":
            _row = {"titulo": ws["A1"].value, "descri": [ws["A2"].value,ws["B2"].value,ws["C2"].value,ws["D2"].value,ws["E2"].value,ws["F2"].value,ws["G2"].value]} 
                        
        return _row

    def get_arch(self, titulo): 
            
        userID = self.request.session['user_list'][0]
        
        form = FicheroForm(self.request.POST, self.request.FILES)
        
        context = {'archivo': '', 'is_valid': False}
        
        if form.is_valid():
            is_valid = True
            files = self.request.FILES.getlist('pic')
            
            for arch in files:
                if Path(str(arch)).suffix in self.extfile:
                    arch_ins = Fichero(pic = arch, titulo = titulo, user_id = userID)
                    arch_ins.save() 
            
            dictpdfs = super().get_fila()
            
            pdfs = []
            for key,value in dictpdfs.items():
                for ki,vi in value.items():
                    if ki == 'pic':
                        pdfr = str(vi).replace('_', ' ')
                        pdfs.append(f'media/{pdfr}')
            
            if len(pdfs):
                pdfsort = sorted(pdfs)
            else:   
                return context
                
            if self.request.POST['archi'] == 'nomunir':  
                fusionador = PdfMerger()
                for pdf in pdfsort:
                    pdfr = str(pdf).replace(' ', '_')
                    fusionador.append(pdfr)  
             
                output = f'media/Archivos/{self.request.POST["nomunir"]}.pdf'
                fusionador.write(output)
                fusionador.close() 
                
                context = {'archivo': output, 'is_valid': True}
            
            elif self.request.POST['archi'] == 'convexcel':
                 
                filepdf = str(pdfsort[0]).replace(' ', '_') 
                x = tabula.read_pdf(filepdf, pages='all', multiple_tables=True)
                for i in  range(len(x)):   #x values in list []
                    df = pd.DataFrame(x[i])
                output = f'media/Archivos/{titulo}.xlsx'
                df.to_excel(output.format(i), header=True, index = True)
                 
            elif self.request.POST['archi'] == 'convedacoop': 
                arch = str(pdfsort[0]).replace(' ', '_')
                fileexcel = f"{arch}"
                
                # Read the Excel file                            
                df = pd.read_excel(fileexcel, sheet_name='Hoja1')
                # Convert the DataFrame to JSON
                json_data = df.to_json(orient='records')
                
                # Write the JSON data to a file
                dir = 'media/Archivos/' 
                
                file_name = Path(fileexcel).stem
                file_name = f"{file_name}.json" 
                with open(os.path.join(dir, file_name), 'w', encoding='ISO-8859-1') as file:
                    json.dump(json_data, file) 
                
                # Define variable to load the dataframe
                excel_dataframe = openpyxl.load_workbook(fileexcel)
                
                # Define variable to read sheet
                dataframe = excel_dataframe.active  

                json_data = []
                titulos = []
                # Iterate the loop to read the cell values
                json_data2 = {"banco": [ws["A1"].value]}
                json_data3 = {"lapso": [ws["A2"].value]}
                json_data4 = {"titulo": [ws["B3"].value,ws["C3"].value,ws["D3"].value,ws["E3"].value,ws["F3"].value,ws["G3"].value,ws["H3"].value]} 
                       
                for row in range(3, dataframe.max_row):                   
                    _row = {}
                    j=0
                    for col in dataframe.iter_cols(1, dataframe.max_column):
                       
                        _row[titulos[j]] = valor
                        
                        j += 1
                             
                    json_data.append(_row)
                
                dictdata = [json_data, json_data2, json_data3, json_data4]
                
                dataobj = super().fnt_insert(101, {"archivo": dictdata})
                #print(context)
                print('max_row',dataframe.max_row)    
                context = {"archivo": f'{dir}{file_name}', 'is_valid': True}
                
                return context 
            
            elif self.request.POST['archi'] == 'convedacoop1': 
                  
                # Read the Excel file
                arch = str(pdfsort[0]).replace(' ', '_')
                fileexcel = f"media/Archivos/{arch}"
                json_data2 = self.filaexcel(fileexcel, 'Hoja2')
                json_data3 = self.filaexcel(fileexcel, 'Hoja3')
                json_data4 = self.filaexcel(fileexcel, 'Hoja4')
                
                df = pd.read_excel(fileexcel, sheet_name='Hoja1')
                # Convert the DataFrame to JSON
                json_data = df.to_json(orient='records')
                
                # Write the JSON data to a file
                dir = 'media/Archivos/' 
                file_name = Path(fileexcel).stem
                file_name = f"{file_name}.json"
                with open(os.path.join(dir, file_name), 'w', encoding='ISO-8859-1') as file:
                    json.dump(json_data, file) 
                json_data = json_data.replace('[','')
                json_data = json_data.replace(']','')
                thisdata = {}
                
                print(json_data)
                #thisdata = {"archivo": json_data, "json_data2": json_data2, "json_data3": json_data3, "json_data4": json_data4}
                
                #dataobj = super().fnt_insert(101, {"archivo": json_data})
                    
                context = {"archivo": f'{dir}{file_name}', 'is_valid': True}
                  
        return context 
		
class MySuper(object):
    def __init__(self,a,b):
        self.a = a
        self.b = b

class MySub(MySuper):
    def __init__(self,a,b,c):
        self.c = c
        super().__init__(a,b)
'''
my_sub = MySub(42,'chickenman','chuco')
print(my_sub.a)
print(my_sub.b)
print(my_sub.c)
		
testdict = {"archivo": "[{\"Nro.\":1,\"Fecha\":\"02-05-2024\",\"Referencia\":135028952.0,\"C\\u00f3digo\":\"NC\",\"Concepto\":\"PAGO MOVIL P2C O.BANCO VAR.\",\"D\\u00e9bito\":0.0,\"Cr\\u00e9dito\":364.7,\"Saldo\":46801.82},{\"Nro.\":876,\"Fecha\":\"21-05-2024\",\"Referencia\":54635763.0,\"C\\u00f3digo\":\"ND\",\"Concepto\":\"COM PAGO MOVIL P2C O.BANCO VAR\",\"D\\u00e9bito\":2.19,\"Cr\\u00e9dito\":0.0,\"Saldo\":5913.86},{\"Nro.\":877,\"Fecha\":\"21-05-2024\",\"Referencia\":54635763.0,\"C\\u00f3digo\":\"ND\",\"Concepto\":\"P2C COM.GASTOS ADMINIST.\",\"D\\u00e9bito\":0.2,\"Cr\\u00e9dito\":0.0,\"Saldo\":5913.66}]", "json_data2": {"titulo": ["Banco"], "descri": ["Bicentenario Banco Universal"]}, "json_data3": {"titulo": ["Movimientos"], "descri": ["01-05-2024 al 22-05-2024"]}, "json_data4": {"titulo": "T\u00edtulos", "descri": ["Nro.", "Fecha", "Referencia", "C\u00f3digo", "Concepto", "D\u00e9bito", "Cr\u00e9dito", "Saldo"]}, "is_valid": True}
for key,value in testdict.items():
    print(key,value)
'''    
    
    
		
			
		
		
