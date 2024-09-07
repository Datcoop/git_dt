import json
import psycopg2
from django.db import connection
import datetime
from datetime import datetime
import time
from datetime import timedelta
from datetime import date
import os 
import os.path
from pathlib import Path

import requests
from bs4 import BeautifulSoup

from .forms import FileFieldForm, FicheroForm
from .models import Fichero

import openpyxl
import tabula
from tabula import read_pdf
import pandas as pd
import PyPDF2
from PyPDF2 import PdfReader , PdfWriter, PdfMerger
#import numpy as np
import csv
import glob
from openpyxl import Workbook
from tabulate import tabulate

class Fntsql:
    def __init__(self, cmdtext):
        self.cmdtext = cmdtext
            
    def dict_selectdattbl(self, funcion):
  
        comi = "'"
        
        commandtext = f'select * from {funcion} ({self.cmdtext});'
        
        rawData = self.exe_sp(commandtext)
        
        this_list = []
        for i, fila in enumerate(rawData):
            this_list.append(rawData[i][0])
        
        return this_list
        
    def dict_ins_filas(self, num_tbl, nominsert, xcamposinsert, pcampos, jsontextvalparam):  
        comi = "'"
        commandtext = f'select * from ins_filas({num_tbl},{comi}{nominsert}{comi},{comi}{xcamposinsert}{comi},{comi}{pcampos}{comi},{comi}{jsontextvalparam}{comi});'     
        
        return self.exe_sp(commandtext)                
    
    def dict_sp_camposinsert(self, num_tblu):  
        commandtext = f'select * from sp_camposinsert({num_tblu});'    
        
        return self.exe_sp(commandtext) 
            
    def fnt_insert(self, num_tbl, dictpart):
      eserr = False
      try:
       
        rawData = self.dict_sp_camposinsert(num_tbl) 
        camposinsert = rawData[0][0]['camposinsert']
        pcampos = rawData[0][0]['pcampos']
        
        listcamposinsert = camposinsert.split(', ') 
            
        listpcampos = pcampos.split(', ')
        
        xnominsert = ''
        pcampos = ''
        xcamposinsert = ''
        for inx in listcamposinsert:
            xnominsert += f'{listpcampos[int(inx)].split(" ")[0]}, '
            pcampos += f'{listpcampos[int(inx)]}, '
            xcamposinsert += f'x.{listpcampos[int(inx)].split(" ")[0]}, '
        
        xnominsert += '.'    
        nominsert = xnominsert.replace(', .', '')   
            
        pcampos += '.'    
        pcampos = pcampos.replace(', .', '')    
            
        xcamposinsert += '.'    
        xcamposinsert = xcamposinsert.replace(', .', '')
        
        constring = f'[{dictpart}]'
        constring = constring.replace("'",'"')
        
        rawData = self.dict_ins_filas(num_tbl, nominsert, xcamposinsert, pcampos, constring) 
        #print(rawData)
        return rawData
      except Exception as e:
        print(e) 
        eserr = True  
        return eserr 
    
    def dict_sp_camposupdate(self, num_tblu):  
        commandtext = f'select * from sp_camposupdate({num_tblu});'     
        
        return self.exe_sp(commandtext) 
 
    def dict_update_filas(self, jsontextvalparam): 
     
        comi = "'"
        
        commandtext = f'select * from update_filas({jsontextvalparam});'     
        
        return self.exe_sp(commandtext) 
        
    def get_fila(self):
        
        this_list = self.dict_selectodo()
        
        return this_list
            
    def dict_selectuno(self):
    
        comi = "'"
        
        commandtext = f'select * from sp_selectuno ({self.cmdtext});'
        
        rawDatauno = self.exe_sp(commandtext)
    
        for i, fila in enumerate(rawDatauno):
            thisdict = {}
            fldcombos = {}
            thisdict = rawDatauno[i][0]
            for i,vi in thisdict.items():
                if i == 'numreg':
                    fldcombos['numreg']=vi
                if i == 'numpags':
                    fldcombos['numpags']=vi
                if i == 'titulotbl':
                    fldcombos['titulotbl']=vi
                if i == 'row_to_json':
                    fldcombos['row_to_json'] = vi
                    fldcombos['fldselect']=vi['fldselect']
                    fldcombos['fldtitulos']=vi['fldtitulos']
                    fldcombos['relacional']=vi['relacional']
        
        return fldcombos
         
    def dict_selectodo(self,cmdtext):
    
        comi = "'"
        
        commandtext = f'select * from sp_selectodo ({self.cmdtext});'
        
        rawData = self.exe_sp(commandtext)
        
        this_list = []
        for i, fila in enumerate(rawData):
            this_list.append(rawData[i][0])
            
        return this_list
         
    def fnt_selectcount(self,num_tbl):
    
        comi = "'"
        
        commandtext = f'select * from sp_selectcount ({num_tbl});'
        
        numregs = self.exe_sp(commandtext)
        
        return numregs 
    
    def fnt_selectcountwhere(self, num_tbl, iglike): 
    
        comi = "'"
        
        commandtext = f'select * from sp_selectcountwhere({num_tbl},{comi}{iglike}{comi});' 
        
        numregs = self.exe_sp(commandtext)
        
        this_list = []
        for i, fila in enumerate(numregs):
            this_list.append(numregs[i][0])
            
        return this_list
    
    def dict_findnumrow(self, num_tbl, numfila):   
        comi = "'"
        commandtext = f'select * from findnumrow({num_tbl},{numfila});'  
        
        rawData = self.exe_sp(commandtext)
        
        this_list = []
        for i, fila in enumerate(rawData):
            this_list.append(rawData[i][0])
            
        return this_list
    
    def get_titulotbl(self, numtbl):
    
        comi = "'"
        
        commandtext = f'select * from titulotbl ({numtbl});'
        
        rawDatauno = self.exe_sp(commandtext)
        
        return rawDatauno[0][0]
        
#////////////////////////////////////////////          
    def exe_sp(self, commandtext):

        print(commandtext)
    
        try:
            with connection.cursor() as cursor:
            
                cursor.execute(commandtext)
                rawData= cursor.fetchall()
               
                return rawData

        finally:
            cursor.close() 

class Llenar:
    # Constructor
    def __init__(self, num_tbl, ruta, requestpost):
        self.num_tbl = num_tbl
        self.ruta = ruta
        self.requestpost = requestpost
        
    def fnt_dictoperkey(self, valor): 
        inxcol = int(valor.split(' ')[0])
        valor1 = valor.split(' ')[2].replace('%','')
        
        operlg = {"ini": "0 > 0", "Igual a": f"{inxcol} = {valor1}", "Mayor o igual a": f"{inxcol} >= {valor1}", "Mayor que": f"{inxcol} > {valor1}", "Menor o igual a": f"{inxcol} <= {valor1}", "Menor que": f"{inxcol} < {valor1}", "Distinto de": f"{inxcol} <> {valor1}", "Comience en": f"{inxcol} ILIKE {valor1}%", "Contenga a": f"{inxcol} ILIKE %{valor1}%", "Termine en": f"{inxcol} ILIKE %{valor1}", "No comience en": f"{inxcol} NOT_ILIKE {valor1}%", "No contenga a": f"{inxcol} NOT_ILIKE %{valor1}%", "No termine en": f"{inxcol} NOT_ILIKE %{valor1}"}
        
        for key,value in operlg.items():
            if str(value) == str(valor):
                return key
        
    def fnt_dictoperval(self, inxcol, key, valor1, valor2):  
        
        operlg = {"ini": "0 > 0", "Igual a": f"{inxcol} = {valor1}", "Mayor o igual a": f"{inxcol} >= {valor1}", "Mayor que": f"{inxcol} > {valor1}", "Menor o igual a": f"{inxcol} <= {valor1}", "Menor que": f"{inxcol} < {valor1}", "Distinto de": f"{inxcol} <> {valor1}", "Comience en": f"{inxcol} ILIKE {valor1}%", "Contenga a": f"{inxcol} ILIKE %{valor1}%", "Termine en": f"{inxcol} ILIKE %{valor1}", "No comience en": f"{inxcol} NOT_ILIKE {valor1}%", "No contenga a": f"{inxcol} NOT_ILIKE %{valor1}%", "No termine en": f"{inxcol} NOT_ILIKE %{valor1}", "Mayor que y Menor que": f"{inxcol} > {valor1} AND {inxcol} < {valor2}", "Mayor o igual a y Menor o igual a": f"{inxcol} >= {valor1} AND {inxcol} <= {valor2}", "Mayor o igual a y Menor que": f"{inxcol} >= {valor1} AND {inxcol} < {valor2}", "Mayor que y Menor o igual a": f"{inxcol} < {valor1} AND {inxcol} <= {valor2}"}
        
        for i,vi in operlg.items():
            if i == key:
                return vi     
        
    def cmdtextfnt(self):
        
        comi = "'"
            
        pagant = self.requestpost['pagant']
    
        param = self.requestpost['param']
        
        numpags = self.requestpost['numpags']
        
        arrwer = self.requestpost['arrwer']
        
        valwhere = self.requestpost['valwhere']
        
        ascdesc = self.requestpost['ascdesc']
        
        ascdesc = self.fntascdesc(param,ascdesc.lower())
        
        fldord = ascdesc['fldord']
        
        ordpage = ascdesc['ordpage']
        
        ascdesc = ascdesc['ascdesc']
        
        paramfil = self.fntprfil(param, numpags, pagant, fldord, arrwer)
        
        arrparamfil = paramfil['paramfil'].split(',')
        
        pagant = arrparamfil[3] 
        indice = arrparamfil[4] 
        limite = arrparamfil[5] 
        
        valwhere = paramfil['valwhere']
        
        cmdtext = f'{self.num_tbl},{comi}{ascdesc}{comi},{paramfil["paramfil"]},{comi}{valwhere}{comi},{comi}0{comi},{comi}0{comi}' 
        
        cmdtext = {'cmdtext': cmdtext, 'argprm': self.requestpost['argprm'], 'irpag': self.requestpost['irpag'], 'paramfil': paramfil['paramfil'], 'ordpage': ordpage, 'arrprm': param.split('-'), 'valwhere': valwhere, 'indice': indice, 'pagant': pagant, 'limite': limite}
        
        return cmdtext
        
    def fntascdesc(self,param,ascdesc):
        
        comi = "'"
                    
        fldord = self.requestpost['fldord']
        
        arrprm = param.split('-')
        print('arrprm1 =',arrprm)
        dictord = {'sort': 'asc', 'asc': 'desc', 'desc': 'asc', 'reset': 'sort'}
        
        if arrprm[2] == 'reset':
            ascdesc = 'sort'
        
        ordpage = self.requestpost['ordpage']
        
        if arrprm[2] == 'page':
            ordpage = 'ord-0-sort'
            
        obj = {'ascdesc': ascdesc, 'ordpage': ordpage, 'fldord': fldord}
        
        listascdesc = ['sort', 'reset']
        
        if arrprm[0] == 'ord':
            ascdesc = arrprm[2]
            print('xxxascdesc',ascdesc)
            fldord = int(arrprm[1])+1
            sortord = dictord[ascdesc]
            print('xxxsortord',sortord)
            ordpage = f'ord-{fldord}-{sortord}'
            if ascdesc in listascdesc:
                ascdesc = 'asc'
            obj = {'ascdesc': ascdesc, 'ordpage':ordpage, 'fldord': fldord}
        
        return obj
        
    def fntprfil(self, param, numpags, pagant, fldord, arrwer):
        
        comi = "'"  
        
        fldbus = self.requestpost['fldbus']
        accion = self.requestpost['accion']
        limite = self.requestpost['regxpag']
        
        listparamfil = self.requestpost['paramfil'].split(',')
        
        dictparamfil = {'fldbus': listparamfil[0], 'fldord': listparamfil[1], 'accion': listparamfil[2], 'pagant': listparamfil[3], 'indice': listparamfil[4], 'limite': listparamfil[5]}
        
        arrprm = param.split('-')
        
        indice = dictparamfil['indice']
        if int(arrprm[1]) > 1 and arrprm[2] == 'ord':
            indice = arrprm[1]
        
        valwhere = arrwer
            
        if arrprm[2] == 'irpag':
            indice = self.requestpost['irpag']   
            if int(indice) > int(numpags):
                indice = numpags        
        elif arrprm[2] == 'page':
            indice = arrprm[1]           
        elif arrprm[2] == 'pgfin':
            indice = numpags            
        elif arrprm[2] == 'reset':
            indice = 1
            pagant = 1
            fldord = 0      
        if arrprm[0] == 'avan':
            fldbus = int(arrprm[1])+1
            listtextbox = [self.requestpost['boxizq1'],self.requestpost['boxder1']]   
            valwhere += f' AND {self.fnt_dictoperval(fldbus, arrprm[2], listtextbox[0], listtextbox[1])}'
            
        listprm0 = ['ini', 'ord', 'lim', 'avan']
        
        listinxextr = [0, int(numpags)+1]
        
        listprevnext = ['Previous', 'Next']
        
        dictprevnext = {'Previous': -1, 'Next': 1}
        
        if arrprm[2] in listprevnext: 
            inx = indice
            indice = int(indice) + dictprevnext[arrprm[2]]
            if indice in listinxextr:
                indice = inx
            else:
                indice = str(indice)
                pagant = inx
            
        paramfil = f'{fldbus},{fldord},{accion},{pagant},{indice},{limite}' 
        
        dictparamfil = {'paramfil': paramfil, 'valwhere': valwhere}
        
        return dictparamfil

class Util(Fntsql):
    def __init__(self, cmdtext):
        super().__init__(cmdtext)

    

    def buildinputbox(self, inpbox, arrfltg, arrcampinsshow, arrrel, arrtit, arrvaltigg, arrfld, arrvalinp):
                    j=0
                    for tit in range(len(arrfltg)):
                        if arrcampinsshow[j] != 'vacio':
                            tp = arrrel[j].split('u')[0]
                            if tp == '10':
                                arrtbl = arrrel[j].split('u')[1].split('x')
                                for tbl in arrtbl:
                                    cmdtext = f"{tbl},'ASC',1,0,7,3,1,10,'0 > 0','0','0'"
                                    cbotodo = super().dict_selectodo(cmdtext)
                                    sel = f'<div class="form-floating p-1"><select class="form-select bg-light" id="inp-{j}" name="inp-{j}" aria-label="Floating label select">'
                                    if arrvaltigg == 'edit':
                                        valfld = arrvalinp[arrfld[j]]
                                        arrvalinprow = super().dict_findnumrow(tbl,valfld)[0]
                                    seleccionado = ''
                                    for dictcbo in cbotodo:
                                      for key,opt in dictcbo.items():
                                        if key == 'numrows':
                                            numrows = dictcbo["numrows"]
                                        else:
                                            if arrvaltigg == 'edit':
                                                if str(opt) == str(arrvalinprow[arrfld[j]]):
                                                    seleccionado = 'selected'
                                            sel += f'<option value="{numrows}" {seleccionado}>{opt}</option>'
                                    titulo = fldcombos.get_titulotbl(tbl)
                                sel += f'</select><label for="floatingSelect">{titulo}</label></div>'
                                inpbox += sel 
                            elif tp == '11':
                                cmdtext = f"{num_tbl},'ASC',1,0,7,3,1,10,'0 > 0','0','0'"
                                contextodo = super().dict_selectodo(cmdtext)
                                valdatalist = ''
                                if arrvaltigg == 'edit':
                                    valdatalist = arrvalinp[arrfld[j]]
                                sel = f'<div class="col-md p-1"><div class="form-floating"><input list="dtl-{j}" type="number" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtit[j]}" aria-describedby="basic-addon1" value="{valdatalist}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtit[j]}</label></div></div>'
                                sel += f'<datalist id="dtl-{j}">'
                                for dictcbo in contextodo:
                                  for key,opt in dictcbo.items():
                                    if key == 'codigo':
                                        sel += f'<option value="{opt}"></option>'
                                titulo = fldcombos.get_titulotbl(tbl)
                                sel += f'</datalist>'
                                inpbox += sel 
                                    
                            else: 
                                if arrvaltigg == 'edit':
                                    print(arrrel[j], arrtit, j, arrvalinp, arrfld) 
                                    inpbox += self.fnt_dictcontrolhtml(arrrel[j], arrtit, j, arrvalinp[arrfld[j]])
                                else:
                                    tp = arrrel[j] 
                                    valdefault = self.fnt_dattype(tp)
                                    inpbox += self.fnt_dictcontrolhtml(arrrel[j], arrtit, j, valdefault)
                                    print(inpbox) 
                        else:
                            if arrfltg[j] != 'vacio':  
                                if str(arrvaltigg[j]) == '20':
                                    numid = request.session['user_list'][0]
                                    inpbox += self.fnt_dictcontrolhtml('20', arrtit, j, numid)
                                elif str(arrvaltigg[j]) == '21':
                                    dtnow = str(datetime.now())
                                    dtnow = f"{dtnow.split(' ')[0]} 09:00:00"
                                    print("dtnow =",dtnow)
                                    inpbox += self.fnt_dictcontrolhtml('21', arrtit, j, dtnow)
                          
                        j += 1
                        
                    return inpbox
                        
    def insertdata(self,num_tbl,dictdata):
                    
        insvaldolar = super().fnt_insert(num_tbl,dictdata)
    
        return dictdata
   
    def actualizartasa(self,candias): 
        listdatos = candias.split(';')
        fechval = listdatos[0]
        strfechval = f'{fechval} 09:00:00'
        candias=int(listdatos[1])
        strfechval = datetime.fromisoformat(strfechval)+timedelta(days=candias)
        valdolar = float(listdatos[2])       
        valdolar = round(valdolar, 4)
        
        dictval = {"tasa": valdolar, "fecha": strfechval}
                    
        insvaldolar = super().fnt_insert(98,dictval)
    
        return dictval
         
        r = requests.get('https://www.bcv.org.ve/glosario/cambio-oficial')
        s = BeautifulSoup(r.content, 'html.parser')
        
        dictval = {}
        if r.status_code == 200:
        
            listfechval = []
            for text in s.find_all('span',{'class':'date-display-single'}):
                b = text
                listfechval.append(b)
                fechval = str(listfechval[0]) 
                fechval = fechval.split(' content="')
                fechavallocal = fechval[1].split('">')[1].replace('</span>','')
            
                fechval = fechval[1].split('T')[0]
                strfechval = f'{fechval} 09:00:00'
                candias=int(candias)
                strfechval = datetime.fromisoformat(strfechval)+timedelta(days=candias)
            valdolar = []
            for text in s.find_all('div',{'class':'col-sm-6 col-xs-6 centrado'}):
                b = text
                valdolar.append(b)    
            valdolar = str(valdolar[4]).replace(' ', '')
            valdolar = valdolar.split('<strong>')
            valdolar = valdolar[1].split('</strong>')
            valdolar = valdolar[0].replace(',', '.') 
            valdolar = float(valdolar)       
            valdolar = round(valdolar, 4)
        
            dictval = {"tasa": valdolar, "fecha": strfechval}
                    
            insvaldolar = super().fnt_insert(98,dictval)
    
        return dictval
  
    def valhasta(self, num_tbl): 
        dictasa = super().dict_selectdattbl('fnt_selecttbl')[0]
        
        return dictasa
            
    def fnt_dattype(self, key): 
        key = int(key)
        comi = "'"
        fecha_actual = datetime.now()
        if key == 3:
            fecha_actual = date.today()
            
        dattype = {1: 0, 2: f"", 3: f"{fecha_actual}", 4: f"{fecha_actual}", 5: f"{fecha_actual}", 6: f"", 7: f"", 8: 0.00, 9: False}
        
        for i,vi in dattype.items():
            if i == key:
                return vi
        
    def fnt_dictoperdescri(self, numoperdict):
        
        operdescri = {"1": "Igual a,Mayor o igual a,Mayor que,Menor o igual a,Menor que,Mayor que y Menor que,Mayor o igual a y Menor o igual a,Mayor o igual a y Menor que,Mayor que y Menor o igual a,Comience en,Contenga a,Termine en,Distinto de,No comience en,No contenga a,No termine en","2": "Igual a,Comience en,Contenga a,Termine en,Distinto de,No comience en,No contenga a,No termine en","3": "Igual a,Mayor o igual a,Mayor que,Menor o igual a,Menor que,Mayor que y Menor que,Mayor o igual a y Menor o igual a,Mayor o igual a y Menor que,Mayor que y Menor o igual a,Comience en,Contenga a,Termine en,Distinto de,No comience en,No contenga a,No termine en","4": "opciones","5": "Igual a,Comience en,Contenga a,Termine en,Distinto de,No comience en,No contenga a,No termine en","8": "Igual a,Mayor o igual a,Mayor que,Menor o igual a,Menor que,Mayor que y Menor que,Mayor o igual a y Menor o igual a,Mayor o igual a y Menor que,Mayor que y Menor o igual a,Comience en,Contenga a,Termine en,Distinto de,No comience en,No contenga a,No termine en","9": "boolean","6": "jsonb"}
        print(operdescri)
        listoperdescri = operdescri[numoperdict].split(',')
        
        dataobje = {  
            'listoperdescri': listoperdescri
        } 
        return dataobje  
        
    def fntchecked(self,valinp):
        checked = ''
        if valinp == True:
            checked = 'checked'    
        return checked            
        
    def fnt_dictcontrolhtml(self, typefld, arrtitulos, j, valinp):
        comi = "'"
        arrjson = 0
        checked = ''
        if int(typefld) == 6:
            arrjson = 6        
        elif int(typefld) == 9:
            checked = self.fntchecked(valinp)
            
        numtype = {
        '1': f'<div class="col-md p-1"><div class="form-floating"><input type="number" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" value="{valinp}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtitulos[j]}</label></div></div>', 
        '2': f'<div class="col-md p-1"><div class="form-floating"><input type="text" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" value="{valinp}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtitulos[j]}</label></div></div>',
        '3': f'<div class="col-md p-1"><div class="form-floating"><input type="date" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" value="{valinp}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtitulos[j]}</label></div></div>', 
        '4': f'<div class="col-md p-1"><div class="form-floating"><input type="datetime-local" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" value="{valinp}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtitulos[j]}</label></div></div>',  
        '5': f'<div class="col-md p-1"><div class="form-floating"><input type="datetime-local" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" value="{valinp}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtitulos[j]}</label></div></div>', 
        '6': f'{arrjson}', 
        '7': f'<div class="col-md p-1"><div class="form-floating"><textarea rows="10" cols="50" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" style="font-size: 14px; color: #5b5b68";></textarea><label for="inp-{j}">{arrtitulos[j]}:{valinp}</label></div></div>', 
        '8': f'<div class="col-md p-1"><div class="form-floating"><input type="number" step="0.01" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" value="{valinp}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtitulos[j]}</label></div></div>',
        '9': f'<div class="col-md p-1"><div class="form-check form-switch"><input type="hidden" id="inp-{j}" name="inp-{j}" value="{valinp}"><input type="checkbox" id="inp{j}" name="inp{j}" class="form-check-input" onchange="fnt_changesttu({comi}checkstatus{comi},{comi}checkedit{comi},{comi}inp-{j}{comi},{comi}inp{j}{comi})" {checked} ><label class="form-check-label" for="inp-{j}">Estátus</label></div></div>',
        '20': f'<input type="hidden" id="inp-{j}" name="inp-{j}" value="{valinp}">',
        '21': f'<input type="hidden" id="inp-{j}" name="inp-{j}" value="{valinp}">'
        }
        
        #print("str(typefld) =",str(typefld),"..............numtype =",numtype[str(typefld)])
        return numtype[str(typefld)]
            
    def html(self, numestilo):
        style = {
            1: ['class="table table-striped table-bordered table-sm"',''],
            2: ['','style = "background: rgb(10, 10, 10);color: white;"'],
            3: ['','style="text-align: center;'],
            4: ['class="link-secondary link-offset-2-opacity-50-hover"','style="text-align: center;cursor: pointer"']
        }
        
        return style[numestilo]
        
    def tbl(self, num_tbl, pagi, paramfil, arrprm, ordpage, tipuser):
       
        comi = "'"
        
        context = super().dict_selectuno()
        
        arrtitulos = context['fldtitulos'].split(", ")
        
        style = self.html(1)
        
        tabla = f'<table id="tblbase" {style[0]}>'
        
        style = self.html(4)
        
        thead = f'<thead class="table-light"><tr><th>Acción</th><th onclick="llenarfld({comi}/tasks/pages/{num_tbl}/1{comi},{comi}frmparam{comi},{comi}ord-0-reset{comi})" {style[0]}{style[1]}>N°</th>'
        tfoot = '<tfoot class="table-light"><tr><td>Acción</td><td>N°</td>'
        
        logfirst = arrprm[0]
        
        dictorden = {'sort': '<i class="fa fa-sort text-black-50 p-2" aria-hidden="true"></i>', 'asc': '<i class="fa fa-sort-amount-asc text-primary p-2" aria-hidden="true"></i>', 'desc': '<i class="fa fa-sort-amount-desc text-primary p-2" aria-hidden="true"></i>'}
        
        listlogfirst = ['ini','page','irpag','avan']
        
        arrordpage = ordpage.split('-')
        
        inx = 0
        for titulo in arrtitulos:
            ordfa = '<i class="fa fa-sort text-black-50 p-2" aria-hidden="true"></i>'
                
            ascdesc = 'sort'
            ordpage = f'ord-{inx}-{ascdesc}'
            if inx+1 == int(arrordpage[1]):
                ascdesc = arrordpage[2]
                ordfa = dictorden[ascdesc]
                print(ascdesc,ordfa)
            ordpage = f'ord-{inx}-{ascdesc}'
            thead += f'<th onclick="llenarfld({comi}/tasks/pages/{num_tbl}/1{comi},{comi}frmparam{comi},{comi}{ordpage}{comi})" {style[0]} {style[1]}>{ordfa}{titulo}</th>'
            inx += 1
            tfoot += f'<td>{titulo}</td>'
            
        thead +=f'</tr></thead>'
        
        contextodo = super().dict_selectodo(self.cmdtext) 
        
        if contextodo:
          tr = ''
          for this_dict in contextodo:
            td = ''
            for key,value in this_dict.items(): 
                c = this_dict['numrows']
                td += f'<td>{value}</td>'
            paramedit = f'style="cursor: pointer" onmousedown="filinp({comi}valup{comi}, {c})" onmouseup="llenarfld({comi}/tasks/pages/{num_tbl}/8{comi},{comi}frmparam{comi},{comi}edit-{c}-edit{comi})"><i class="fa fa-edit p-1 bg-primary text-light rounded"></i></span>' 
            if tipuser != 1 and num_tbl == 98:   
                paramedit = ''
            tr += f'<tr><td class="text-center"><span data-bs-toggle="offcanvas" data-bs-target="#offcanvasRight" aria-controls="offcanvasRight" {paramedit}</td>{td}</tr>'
        else:
          tr = f'<tr><td colspan="{len(arrtitulos)+2}">No hay datos para mostrar</td></tr>'
         
            
        tbody = f'<tbody>{tr}</tbody>'
          
        numreg = context['numreg']
        
        numpags = context['numpags']
        
        tfoot += '</tr></tfoot>'
            
        tabla += f'{thead}{tbody}{tfoot}</table>'  
        
        paramfil = paramfil.split(',')
        pagant = int(paramfil[3])   
        indice = int(paramfil[4])    
        limite = int(paramfil[5])  
              
        dataobje = {  
            'tabla': tabla,
            'idres': 'listado',
            'url': f'/tasks/pages/{num_tbl}/2',
            'recordsTotal': numreg,
            'numpags': numpags,
            'indice': indice,
            'limite': limite,
            'pagedat': pagi,
        }
        
        return dataobje 

    def paginacion(self, paramfil, numtbl, arrtitulos, numreg, numpags, argprm, irpag):
    #'['1', '0', '5', '1', '1', '10']'
     try:
      if int(numpags) > 0:
        comi = "'"
        paramfil = paramfil.split(',')
    
        pagant = int(paramfil[3])    
        indice = int(paramfil[4])   
        limite = int(paramfil[5])
        valder = indice*limite
        if indice == int(numpags):
            valder = int(numreg)
        
        pagi = f'<div class="col-3"><p style="font-size: 13px">(Mostrando {(indice-1)*limite+1} al {valder} de {numreg})</p></div><div class="col-7"><ul class="pagination pg-primary" style="cursor: pointer;"><li class="page-item" onclick="llenarfld({comi}/tasks/pages/{numtbl}/1{comi},{comi}frmparam{comi},{comi}page-{indice}-Previous{comi})"><a class="page-link" style="background-color: rgb(31, 36, 46);color: white;"><i class="fas fa-angle-double-left"></i></a></li>'
           
        active = ''
        for i in range(int(numpags)):
            inx = i + 1 
                
            if inx < 8:
                if inx == indice:
                    active = 'active'
                pagi += f'<li class="page-item {active}" onclick="llenarfld({comi}/tasks/pages/{numtbl}/1{comi},{comi}frmparam{comi},{comi}page-{inx}-page{comi})"><a class="page-link">{inx}</a></li>'
            elif inx == 8:
                if indice == 8:
                    active = 'active'
                pagi += f'<li class="page-item {active}" onclick="llenarfld({comi}/tasks/pages/{numtbl}/1{comi},{comi}frmparam{comi},{comi}page-{inx}-page{comi})"><a class="page-link">{inx}</a></li>'
                if numpags == 8:
                    break
            elif indice > 8 and indice < numpags:
                pagi += f'<li class="page-item"><a class="page-link">...</a></li><li class="page-item active" onclick="llenarfld({comi}/tasks/pages/{numtbl}/1{comi},{comi}frmparam{comi},{comi}page-{indice}-page{comi})"><a class="page-link" >{indice}</a></li><li class="page-item"><a class="page-link">...</a></li><li class="page-item" onclick="llenarfld({comi}/tasks/pages/{numtbl}/1{comi},{comi}frmparam{comi},{comi}page-{numpags}-pgfin{comi})"><a class="page-link">{numpags}</a></li>'
                break
            else:
                if indice == numpags:
                    active = 'active'
                pagi += f'<li class="page-item"><a class="page-link">...</a></li><li class="page-item"><a class="page-link">...</a></li><li class="page-item {active}" onclick="llenarfld({comi}/tasks/pages/{numtbl}/1{comi},{comi}frmparam{comi},{comi}page-{numpags}-pgfin{comi})"><a class="page-link">{numpags}</a></li>'
                break
            active = ''
          
        pagi += f'<li class="page-item" onclick="llenarfld({comi}/tasks/pages/{numtbl}/1{comi},{comi}frmparam{comi},{comi}page-{indice}-Next{comi})"><a class="page-link" style="background-color: rgb(31, 36, 46);color: white;"><i class="fas fa-angle-double-right"></i></a></li></ul></div><div class="col-2 m-auto"><div class="input-group mb-3"><input type="text" id="irpagnum" name="irpagnum" class="form-control" placeholder="Ir a pág..." aria-label="Ir a pág..." aria-describedby="ir-pag" onchange="fillbox({comi}irpagnum{comi},{comi}irpag{comi})"><span class="input-group-text" id="ir-pag"><i class="fa fa-search" style="cursor: pointer;" onclick="llenarfld({comi}/tasks/pages/{numtbl}/1{comi},{comi}frmparam{comi},{comi}page-{irpag}-irpag{comi})"></i></span></div></div>'
    
      else:
        pagi = '' 
  
      return pagi

     except Exception as e:
      print(e)
      
class Archivo(Fntsql): 
    # Constructor
    def __init__(self, request, cmdtext, extfile):
        self.request = request
        self.cmdtext = cmdtext
        self.extfile = extfile
        super().__init__(cmdtext)
    
    def fnt_chkisexcel(self, excel_file_path,arch):
        
        wb = openpyxl.load_workbook(excel_file_path)
                # Define variable to read sheet
        hojas = wb.sheetnames
        ws = wb[hojas[0]] 

        listfldselectall = [ws["A2"].value, ws["B2"].value, ws["C2"].value, ws["D2"].value, ws["E2"].value, ws["F2"].value, ws["G2"].value]
        
        listfldselect = []
        for fld in listfldselectall:
            if fld is not None:
                listfldselect.append(fld)
        
        #dictnumtbl = {'listbancos': 100, '': 101, 142]
        namefile = arch.split('.')[0]
                                      
        dictexcel = {
                 100: {'col': ['A','B'], 
                    'fields': ['codigo','nombre']},
                 101: {
                       'col': ['A','B','C','D','E','F','G'],
                    'fields': ['fecha','referencia','codigo','descripcion','debito','credito','saldo']}, 
                 142: {'col': ['A','B'], 
                    'fields': ['codigo','nombre']}
                    } 
        
        if listfldselect == dictexcel[101]["fields"]:
            isvalid = True 
            dataobj = dictexcel[101] 
            num_tblu = 101
        elif listfldselect == dictexcel[142]["fields"]:
            isvalid = True 
            dataobj = dictexcel[142]
            num_tblu = 142
        else:
            isvalid = False 
            num_tblu = 0
            dataobj = {}
        
        dataobj['isvalid'] = isvalid
        
        dataobj['num_tblu'] = num_tblu
        
        return dataobj    
    
    def pdf_to_csv(self, pdf_file_path, csv_file_path, path):
        # Read PDF file
        print(datetime.now(), 'Leyendo el archivo PDF')
        tables = tabula.read_pdf(pdf_file_path, pages='all')
        csv_files = f'{path}libro.csv'
        if tables:
            print(datetime.now(), 'Convirtiendo PDF a CSV')
        # Write each table to a only sheet in the Excel file
            tabula.convert_into(pdf_file_path, csv_files, output_format='csv', pages='all')
            print(datetime.now(), 'Convirtiendo PDF a CSV')
            with open(csv_files, newline='') as csvfile:
                reader = csv.DictReader(csvfile)
                  
            return csv_file_path
            
        else:
        
          return False  
    
    def csv_to_excel(self, csv_file_path, excel_file_path, path):
        # Read PDF file
        print(datetime.now(), 'Leyendo el archivo PDF')
        df = pd.read_csv(csv_file_path)
        # Writing dataframe to an excel file
        tables = df.to_excel(excel_file_path, index=False) 
        
        if tables:       
            return excel_file_path            
        else:        
          return False 
    
    def csv_to_json(self, reader, num_tbl):
        # Read PDF file
      print('reader') 
      try:   
        nomcol = {101: ['Fecha', 'Referencia', 'Código', 'Descripción', 'Débito', 'Crédito', 'Saldo']}
        
        dictvalcol = {101: 'Referencia'}
        
        valcol = dictvalcol[num_tbl]
        
        listnomcol = nomcol[num_tbl]
        
        with open(reader, newline='') as csvfile:          
                reader = csv.DictReader(csvfile)
                dictreader = {}
                f = 1
                for row in reader:
                    dictrow = {}
                    for j in range(len(listnomcol)):
                       dictrow[j] = (row[listnomcol[j]])
                    if len(dictrow[1]) > 0 and dictrow[1] != listnomcol[1]:  
                        dictreader[f] = dictrow
                        f += 1
                  
        return dictreader
      except Exception as e:
        print(e)
              
    def delfile(self,filespath):        
        files = os.listdir(filespath)
        extfile = ['.pdf', '.xls', '.xlsx', '.tif', '.tiff', '.bmp', '.jpg', '.jpeg', '.gif', '.png', '.eps', '.json', '.csv']
    
        for file in files:
            if Path(file).suffix in extfile:
                fileruta = f'{filespath}{file}'
                os.remove(fileruta) # delete file based on their name or suffix  
        
    def get_arch(self, request, titulo, num_tbl): 
        print('vvvvv =',request)
        userID = request.session['user_list'][0]
        print(request.FILES)
        
        PATH = 'datacoop_web'
        if not os.path.exists(PATH):
            os.makedirs(PATH)
    
        form = FicheroForm(request.POST, request.FILES)
        
        context = {'archivo': '', 'is_valid': False}
        
        if form.is_valid():
            is_valid = True
            files = request.FILES.getlist('pic')
            print('files =',files)
            
            for filename, file in request.FILES.items():
                name = request.FILES[filename].name
                print(name)
            
            pdfs = []  
            f = 1  
            for arch in files:
                print('arch =',arch)
                iglike = f"1 = Archivos/{arch}"
                print('iglike =',iglike)
                if Path(str(arch)).suffix in self.extfile:
                    arch_ins = Fichero(pic = arch, titulo = titulo, user_id = userID)
                    numreg = super().fnt_selectcountwhere (100,iglike);
                    print('numreg =',len(numreg),request.POST['archi'])
                    
                    notnumreg = True
                    if not len(numreg):
                        arch = str(arch)
                        arch_ins.save()
                        arch = arch.replace('(','')
                        arch = arch.replace(')','')
                        arch = arch.replace(' ','_')
                        pdfs.append(f'media/Archivos/{arch}')
            if len(pdfs):
                pdfsort = sorted(pdfs)
            else:   
                return context
                
            if request.POST['archi'] == 'nomunir':  
              
                fusionador = PdfMerger()
                for pdf in pdfsort:
                    pdfr = str(pdf).replace(' ', '_')
                    fusionador.append(pdfr)  
                output = f'media/Archivos/{titulo}.pdf'
                fusionador.write(output)
                fusionador.close() 
                context = {'archivo': output, 'is_valid': True}
                
                return context
                      
            elif request.POST['archi'] == 'convecsv':
                print('pdfsort',pdfsort) 
                pdf_file_path = str(pdfsort[0]).replace(' ', '_') 
                output = f'media/Archivos/{titulo}.csv'
                # Define variable to read sheet
                outputcsv = self.pdf_to_csv(pdf_file_path, output, 'media/Archivos/')
                
                if outputcsv: 
                    print(datetime.now(), 'Listo')
                    
                    headers = ['Fecha','Referencia','Código','Descripción','Débito','Crédito','Saldo']
                 
                    context = {"archivo": outputcsv, 'is_valid': True}
                else: 
                    print(datetime.now(), 'Listo')
                 
                    context = {"archivo": outputcsv, 'is_valid': False}
                
                return context 
                      
            elif request.POST['archi'] == 'convexcel':
                print('pdfsort',pdfsort) 
                pdf_file_path = str(pdfsort[0]).replace(' ', '_') 
                output = f'media/Archivos/{titulo}.xlsx'
                # Define variable to read sheet
                outputexcel = self.pdf_to_excel(pdf_file_path, output, 'media/Archivos/')
                
                if outputexcel: 
                    print(datetime.now(), 'Listo')
                    
                    headers = ['Fecha','Referencia','Código','Descripción','Débito','Crédito','Saldo']
                 
                    context = {"archivo": outputexcel, 'is_valid': True}
                else: 
                    print(datetime.now(), 'Listo')
                 
                    context = {"archivo": outputexcel, 'is_valid': False}
                
                return context 
                 
            elif request.POST['archi'] == 'convedacoop':            
                arch = str(pdfsort[0]).replace(' ', '_')
                csv_file_path = arch
                num_tbl = 101
                filejson = self.csv_to_json(csv_file_path, num_tbl)
                print(filejson)     
                json_file_path = f'{PATH}/movimientos_{datetime.now()}.json'
                
                with open(json_file_path, "w") as fp:
                    json.dump(filejson, fp) 
                
                #insvaldolar = super().fnt_insert(num_tblu,_row) 
                fileToWrite = open(json_file_path, "a")
    
                if json_file_path: 
                    print(datetime.now(), 'Listo')
                 
                    context = {"archivo": json_file_path, 'is_valid': True}
                else: 
                    print(datetime.now(), 'Listo')
                 
                    context = {"archivo": json_file_path, 'is_valid': False}
                
                return context 


      
        
