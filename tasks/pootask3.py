from math import pi
from math import sqrt
import sys
import random
import math
import psycopg2
import json
import openpyxl
import tabula
from tabula import read_pdf
import pandas as pd
#import numpy as np
import csv
from xlwt import Workbook 
from tabulate import tabulate
from django.db import connection
from datetime import datetime
import os 
import PyPDF2
from PyPDF2 import PdfReader , PdfWriter, PdfMerger
from gallery.models import *
from gallery.forms import *
from django.views.generic.edit import FormView
from .forms import FileFieldForm, FicheroForm
from .models import Fichero
from django.contrib.auth.models import User
from pathlib import Path
from .models import Cooperativas, Socios

class Coffee:
    # Constructor
    def __init__(self, name, price):
        self.name = name
        self.price = float(price)
    def check_budget(self, budget):
        # Check if the budget is valid
        if not isinstance(budget, (int, float)):
            print('Enter float or int')
            exit()
            if budget < 0: 
                print('Sorry you don\'t have money') 
                exit() 
    def get_change(self, budget):
        return budget - self.price
        
    def sell(self, budget):
        self.check_budget(budget)
        if budget >= self.price:
            print(f'You can buy the {self.name} coffee')
            if budget == self.price:
                print('It\'s complete')
            else:
                print(f'Here is your change {self.get_change(budget)}$')

                exit('Thanks for your transaction')
                        
class Cookie:
	# Constructor
	def __init__(self, name, shape, chips='Chocolate'):
		# Instance attributes
		self.name = name
		self.shape = shape
		self.chips = chips

	# The object is passing itself as a parameter
	def bake(self):
		print(f'This {self.name}, is being baked with the shape {self.shape} and chips of {self.chips}')
		print('Enjoy your cookie!')

class Shape:
	def __init__(self, side1, side2):
		self.side1 = side1
		self.side2 = side2

	def get_area(self):
		return self.side1 * self.side2

	def __str__(self):
		return f'The area of this {self.__class__.__name__} is: {self.get_area()}'
 
class Rectangle(Shape): # Superclass in Parenthesis
	pass
 
class Square(Rectangle):
	def __init__(self, side):
		super().__init__(side, side)
		
class Triangle(Rectangle):
	def __init__(self, base, height):
		super().__init__(base, height)
 
	def get_area(self):
		area = super().get_area()
		return area / 2
 
class Circle(Shape):
	def __init__(self, radius):
		self.radius = radius
 
	def get_area(self):
		return pi * (self.radius ** 2)
 
class Hexagon(Rectangle):
	def __init__(self, side1):
		self.side1 = side1
	
	def get_area(self):
		return (3 * sqrt(3) * self.side1 ** 2) / 2

class Dictoperador:

    def __init__(self):
        self.dicexp = {'1': 'tblpdf', '2': 'tblexcel', '3': 'tblcsv', '4': 'fnt_bucle()'}
            
    def fnt_dattype(self, key): 
        comi = "'"
        fecha_actual = date.today()
        dattype = {"bigint": 0, "integer": 0, "character varying": f"{comi}0{comi}", "date": f"{comi}{fecha_actual}{comi}", "double precision": 0.00}
        
        for i,vi in dattype.items():
            if i == key:
                return vi
        
    def fnt_dictoperdescri(self, numoperdict):
        
        operdescri = {"1": "Igual a,Mayor o igual a,Mayor que,Menor o igual a,Menor que,Mayor que y Menor que,Mayor o igual a y Menor o igual a,Mayor o igual a y Menor que,Mayor que y Menor o igual a,Comience en,Contenga a,Termine en,Distinto de,No comience en,No contenga a,No termine en","2": "Igual a,Comience en,Contenga a,Termine en,Distinto de,No comience en,No contenga a,No termine en","3": "Igual a,Mayor o igual a,Mayor que,Menor o igual a,Menor que,Mayor que y Menor que,Mayor o igual a y Menor o igual a,Mayor o igual a y Menor que,Mayor que y Menor o igual a,Comience en,Contenga a,Termine en,Distinto de,No comience en,No contenga a,No termine en","4": "opciones","5": "Igual a,Comience en,Contenga a,Termine en,Distinto de,No comience en,No contenga a,No termine en","8": "Igual a,Mayor o igual a,Mayor que,Menor o igual a,Menor que,Mayor que y Menor que,Mayor o igual a y Menor o igual a,Mayor o igual a y Menor que,Mayor que y Menor o igual a,Comience en,Contenga a,Termine en,Distinto de,No comience en,No contenga a,No termine en","9": "boolean","6": "jsonb"}
        
        listoperdescri = operdescri[numoperdict].split(',')
        
        dataobje = {  
            'listoperdescri': listoperdescri
        } 
        return dataobje
        
    def fnt_dictopersymbol(self, numoperdict): 
    
        opersymbol = {"1": "0,=,>=,>,<=,<,>y<,>=y<=,>=y<,>y<=,y%s,%sy%s,%sy,y<>,y%n,%ny%n,%ny","2": "0,=,y%s,%sy%s,%y,y<>,y%n,%ny%n,%ny","3": "0,=,>=,>,<=,<,>y<,>=y<=,>=y<,>y<=,y%s,%sy%s,%sy,y<>,y%n,%ny%n,%ny","4": "opt"}
        
        return opersymbol[numoperdict]
        
    def fnt_dictoperlg(self, key): 
        
        operlg = {"Igual a": "=", "Mayor o igual a": ">=", "Mayor que": ">", "Menor o igual a": "<=", "Menor que": "<", "Distinto de": "<>", "Comience en": "y%s", "Contenga a": "%sy%s", "Termine en": "%sy", "No comience en": "y%n", "No contenga a": "%ny%n", "No termine en": "%ny", "Mayor que y Menor que": ">y<", "Mayor o igual a y Menor o igual a": ">=y<=", "Mayor o igual a y Menor que": ">=y<", "Mayor que y Menor o igual a": ">y<="}
        
        for i,vi in operlg.items():
            if i == key:
                return vi
        
    def fnt_dictoperval(self, inxcol, key, valor1, valor2): 
        
        operlg = {"Igual a": f"{inxcol} = {valor1}", "Mayor o igual a": f"{inxcol} >= {valor1}", "Mayor que": f"{inxcol} > {valor1}", "Menor o igual a": f"{inxcol} <= {valor1}", "Menor que": f"{inxcol} < {valor1}", "Distinto de": f"{inxcol} <> {valor1}", "Comience en": f"{inxcol} ILIKE {valor1}%", "Contenga a": f"{inxcol} ILIKE %{valor1}%", "Termine en": f"{inxcol} ILIKE %{valor1}", "No comience en": f"{inxcol} NOT_ILIKE {valor1}%", "No contenga a": f"{inxcol} NOT_ILIKE %{valor1}%", "No termine en": f"{inxcol} NOT_ILIKE %{valor1}", "Mayor que y Menor que": "{inxcol} > {valor1} AND {inxcol} < {valor2}", "Mayor o igual a y Menor o igual a": "{inxcol} >= {valor1} AND {inxcol} <= {valor2}", "Mayor o igual a y Menor que": "{inxcol} >= {valor1} AND {inxcol} < {valor2}", "Mayor que y Menor o igual a": "{inxcol} < {valor1} AND {inxcol} <= {valor2}"}
        
        for i,vi in operlg.items():
            if i == key:
                return vi               
        
    def fnt_dictdirec(self, key): 
        direc = {'0': 'fa fa-sort', 'ASC': 'fa fa-sort-amount-up', 'DESC': 'fa fa-sort-amount-down'}
        
        for i,vi in direc.items():
            if i == key:
                return vi   
        
    def fntchecked(self,valinp):
        checked = ''
        if valinp == True:
            checked = 'checked'    
        return checked            
        
    def fnt_dictcontrolhtml(self, typefld, arrtitulos, j, valinp):
        comi = "'"
        arrjson = 0
        checked = ''
        if int(typefld) == 6:
            arrjson = 6        
        elif int(typefld) == 9:
            checked = self.fntchecked(valinp)
            
        numtype = {
        '1': f'<div class="col-md p-1"><div class="form-floating"><input type="number" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" value="{valinp}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtitulos[j]}</label></div></div>', 
        '2': f'<div class="col-md p-1"><div class="form-floating"><input type="text" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" value="{valinp}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtitulos[j]}</label></div></div>',
        '3': f'<div class="col-md p-1"><div class="form-floating"><input type="date" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" value="{valinp}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtitulos[j]}</label></div></div>', 
        '4': f'<div class="col-md p-1"><div class="form-floating"><input type="datetime-local" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" value="{valinp}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtitulos[j]}</label></div></div>', 
        '6': f'{arrjson}', 
        '7': f'<div class="col-md p-1"><div class="form-floating"><textarea rows="10" cols="50" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" style="font-size: 14px; color: #5b5b68";></textarea><label for="inp-{j}">{arrtitulos[j]}:{valinp}</label></div></div>', 
        '8': f'<div class="col-md p-1"><div class="form-floating"><input type="number" step="0.01" class="form-control bg-light" id="inp-{j}" name="inp-{j}" placeholder="{arrtitulos[j]}" aria-describedby="basic-addon1" value="{valinp}" style="font-size: 14px; color: #5b5b68"><label for="inp-{j}">{arrtitulos[j]}</label></div></div>',
        '9': f'<div class="col-md p-1"><div class="form-check form-switch"><input type="hidden" id="inp-{j}" name="inp-{j}" value="{valinp}"><input type="checkbox" id="inp{j}" name="inp{j}" class="form-check-input" onchange="fnt_changesttu({comi}checkstatus{comi},{comi}checkedit{comi},{comi}inp-{j}{comi},{comi}inp{j}{comi})" {checked} ><label class="form-check-label" for="inp-{j}">Estátus</label></div></div>'
        }
        
        #print("str(typefld) =",str(typefld),"..............numtype =",numtype[str(typefld)])
        return numtype[str(typefld)]
    	
class Diccionario:

    def __init__(self, request, cmdtext):
        """Esta clase se usa para todas las operaciones con diccionarios"""
        self.request = request
        self.cmdtext = cmdtext
    
    def renameKeysToLower(self, orig_dict):
        new_dict = {k.lower(): v for k, v in orig_dict.items()}
        
        return new_dict

    def countreg(self, numtbl):
    
        comi = "'"
        
        commandtext = f'select * from sp_selectcount({numtbl});'
        
        rawDatauno = self.exe_sp(commandtext)
        
        return rawDatauno[0][0]
    
    def get_titulotbl(self, numtbl):
    
        comi = "'"
        
        commandtext = f'select * from titulotbl ({numtbl});'
        
        rawDatauno = self.exe_sp(commandtext)
        
        return rawDatauno[0][0]
        
    def gentbladmin(self, numtbl, ruta):
        comi = "'"        
        tbladmingen = []  
        
        thisdict = self.dict_selectodo()
         
        for key,value in thisdict[0].items():
            if key == 'tablas':
                filatbl = []
                arrnumtbl = value.split(',')
                for numtbl in arrnumtbl:
                    titulo = self.get_titulotbl(int(numtbl))
                    filatbl.append([int(numtbl),titulo,ruta])
        return filatbl
        
    def tbl(self, num_tbl, pagi):
      
        context = self.dict_selectuno()
        
        arrtitulos = context['fldtitulos'].split(", ")
        
        style = self.html(1)
        
        tabla = f'<table id="tblbase" {style[0]}>'
        
        thead = '<thead class="table-light"><tr><th>Acción</th><th>N°</th>'
        tfoot = '<tfoot class="table-light"><tr><td>Acción</td><td>N°</td>'
        for titulo in arrtitulos:
            thead += f'<th>{titulo}</th>'
            tfoot += f'<td>{titulo}</td>'
              
        thead +=f'</tr></thead>'
        
        contextodo = self.dict_selectodo() 
        
        if contextodo:
          tr = ''
          for this_dict in contextodo:
            td = ''
            for key,value in this_dict.items():
                td += f'<td>{value}</td>'
            tr += f'<tr><td></td>{td}</tr>'
        else:
          tr = f'<tr><td colspan="{len(arrtitulos)+2}">No hay datos para mostrar</td></tr>'
          
        tbody = f'<tbody>{tr}</tbody>'
          
        numreg = context['numreg']
        
        numpags = context['numpags']
        
        tfoot += '</tr></tfoot>'
            
        tabla += f'{thead}{tbody}{tfoot}</table>'
        
        paramfil = self.request.session['paramfil']
        paramfil = paramfil.split(',')
        pagant = int(paramfil[3])   
        indice = int(paramfil[4])    
        limite = int(paramfil[5])  
              
        dataobje = {  
            'tabla': tabla,
            'idres': 'listado',
            'url': f'/tasks/pages/{num_tbl}/2',
            'recordsTotal': numreg,
            'numpags': numpags,
            'indice': indice,
            'limite': limite,
            'pagedat': pagi,
        }
        
        return dataobje 
          
    def dict_selectuno(self):
    
        comi = "'"
        
        commandtext = f'select * from sp_selectuno ({self.cmdtext});'
        
        rawDatauno = self.exe_sp(commandtext)
    
        for i, fila in enumerate(rawDatauno):
            thisdict = {}
            fldcombos = {}
            thisdict = rawDatauno[i][0]
            for i,vi in thisdict.items():
                if i == 'numreg':
                    fldcombos['numreg']=vi
                if i == 'numpags':
                    fldcombos['numpags']=vi
                if i == 'titulotbl':
                    fldcombos['titulotbl']=vi
                if i == 'row_to_json':
                    fldcombos['fldselect']=vi['fldselect']
                    fldcombos['fldtitulos']=vi['fldtitulos']
                    fldcombos['relacional']=vi['relacional']
                    fldcombos['tooltipaddreg']=vi['tooltipaddreg']
        
        return fldcombos
         
    def dict_selectodo(self):
    
        comi = "'"
        
        commandtext = f'select * from sp_selectodo ({self.cmdtext});'
        
        rawData = self.exe_sp(commandtext)
        
        this_list = []
        for i, fila in enumerate(rawData):
            this_list.append(rawData[i][0])
            
        return this_list
        
    def get_fila(self):
        
        this_list = self.dict_selectodo()
        
        return this_list[0]
        
    def rawData(self):
        
        rawData = self.dict_selectodo(self.cmdtext)
        
        return rawData
        
    def dict_ins_filas(self, num_tblu, pxcampo, pcampos, jsontextvalparam):  
        comi = "'"
        commandtext = f'select * from ins_filas({num_tblu},{comi}{pxcampo}{comi},{comi}{pcampos}{comi},{comi}{jsontextvalparam}{comi});'     
        
        return self.exe_sp(commandtext)                
    
    def dict_sp_camposinsert(self, num_tblu):  
        commandtext = f'select * from sp_camposinsert({num_tblu});'    
        
        return self.exe_sp(commandtext) 
            
    def fnt_insert(self, num_tbl, dictpart):
      eserr = False
      try:
        
        rawData = self.dict_sp_camposinsert(num_tbl) 
        xnominsert = rawData[0][0]['xcamposinsert']
        pcampos = rawData[0][0]['pcampos']
#{"camposinsert":"codigo","xcamposinsert":"xcodigo","pcampos":"codigo jsonb"} 
        constring = f'[{dictpart}]'
        constring = constring.replace("'",'"')
        
        rawData = self.dict_ins_filas(num_tbl, xnominsert, pcampos, constring) 
        #print(rawData)
        return eserr
      except Exception as e:
        print(e) 
        eserr = True  
        return eserr 
    
    def dict_sp_camposupdate(self, num_tblu):  
        commandtext = f'select * from sp_camposupdate({num_tblu});'     
        
        return self.exe_sp(commandtext) 
 
    def dict_update_filas(self, jsontextvalparam):  
        comi = "'"
        commandtext = f'select * from update_filas({jsontextvalparam});'     
        
        return self.exe_sp(commandtext) 
        
    def fnt_valdefault(self, num_tbl, rel):
        dia = datetime.today().strftime('%d-%m-%Y') 
        diastamp = datetime.now()
        valtipuser = 0
        if int(num_tbl) == 461 and int(rel) == 1:
            valtipuser = self.dict_sp_selectcount(num_tbl)[0][0] + 1
        valdefault = {1: valtipuser, 2: 'vacio', 3: dia, 4: diastamp, 7: 'vacio', 8: 0, 9: False}
        
        return valdefault[int(rel)]
            
    def builddata(self, num_tbl, cmdtext, querydict, esinsert): 
                
        rawData = self.dict_sp_camposupdate(num_tbl) 
        
        for i, filaraw in enumerate(rawData):                
            thisdict = {}         
            thisdict = rawData[i][0]
            j = 0
            for key,value in thisdict.items():
                if key == 'camposupdate':
                    fldsel = value
                elif key == 'xnomupdate':
                    xnomupdate = value
                elif key == 'xcamposinsert':
                    xnominsert = value
                elif key == 'pcampos':
                    pcampos = value
                elif key == 'pcamposout':
                    pcamposout = value
                        
            arrfldsel = fldsel.split(', ')
            
            arrfldselect = self.dict_selectuno['fldselect'].split(", ")  
        
            arrrelacional = self.dict_selectuno['relacional'].split(", ") 
            
            dictval = {}
            c = 0 
            for nomcamp in arrfldsel:
                rel = arrrelacional[c]
                if esinsert:
                    dictval[nomcamp] = self.fnt_valdefault(num_tbl, rel)
                else:
                    ncqdict = f'inp-{c}'
                    dictval[nomcamp] = querydict[ncqdict]
                c += 1 
             
            if len(arrfldselect) == 1 and arrfldselect[0] == 'codigo':
                data = {"codigo": dictval}
            else:
                data = dictval 
                 
            data=json.dumps(data)
            
            dataobj = {"data": data, "xnominsert": xnominsert, "xnomupdate": xnomupdate, "pcampos": pcampos, "pcamposout": pcamposout, "arrfldselect": arrfldselect, "arrrelacional": arrrelacional, "dictval": dictval}
            
            return dataobj
            
    def html(self, numestilo):
        style = {
            1: ['class="table table-striped table-bordered table-sm"',''],
            2: ['','style = "background: rgb(10, 10, 10);color: white;"'],
            3: ['','style="text-align: center;']
        }
        
        return style[numestilo]
        
#////////////////////////////////////////////          
    def exe_sp(self, commandtext):

        print(commandtext)
    
        try:
            with connection.cursor() as cursor:
            
                cursor.execute(commandtext)
                rawData= cursor.fetchall()
               
                return rawData

        finally:
            cursor.close() 
            
class Cattbl(Diccionario):

    def __init__(self, request, cmdtext):
       super().__init__(request, cmdtext)
        
    def infoesquema(self, nomtbl):
        datainfo = self.dict_sp_information_schematblxnombre(nomtbl)
            
        return datainfo
        
    def fnt_numtype(self, typefld):
       
        numtype = {"integer": 1, "bigint": 1, "character varying": 2, "date": 3, "timestamp with time zone": 4, "jsonb": 6, "text": 7, "double precision": 8, "boolean": 9}
        
        for i,vi in numtype.items():
            if str(i) == str(typefld):
                return vi
    
    def dict_sp_information_schematblxnombre(self, nomtbl):
       #numtbl,nomtbl,tittbl
       try:    
        comi = "'"
        commandtext = f'select * from sp_information_schematblxnombre({comi}{nomtbl}{comi});' 

        rawData = super().exe_sp(commandtext) 
       
        campos = {} 
        fldselec = {}  
        fldtitulos = {}        
        relacional = {} 
        camposinsert = {}  
        camposupdate = {} 
        xnomupdate = {}  
        pcampos = {}  
        pcamposout = {}  
        xcamposinsert = {} 
        for i, fila in enumerate(rawData):
            thisdict = {}
            thisdict = rawData[i][0]
            for k,vk in thisdict.items():                   
                if k == 'column_name':
                        nomcampo = vk
                        campos[i] = vk  
                        fldselec[i] = vk   
                        fldtitulos[i] = vk  
                        camposinsert[i] = vk   
                        camposupdate[i] = vk     
                        xnomupdate[i] = f'{vk} = x.{vk}' 
                        pcamposout[i] = f'x.{vk}' 
                        xcamposinsert[i] = f'x.{vk}'                   
                elif k == 'data_type':    
                        relacional[i] = self.fnt_numtype(f'{vk}')  
                        pcampos[i] = f'{nomcampo} {vk}' 
        datacombos =  {                   
            1: campos,
            2: fldselec,
            3: fldtitulos,        
            4: relacional,
            5: camposinsert, 
            6: camposupdate,
            7: xnomupdate,
            8: pcampos,
            9: pcamposout,
           10: xcamposinsert, 
        }  
                     
        camposyy = '' 
        fldselecyy = ''  
        fldtitulosyy = ''        
        relacionalyy = '' 
        camposinsertyy = ''  
        camposupdateyy = '' 
        xnomupdateyy = ''  
        pcamposyy = ''  
        pcamposoutyy = '' 
        xcamposinsertyy = ''     
        
        for num,fldcombos in datacombos.items():
            sizeoff = len(fldcombos)
            fila = 1
            for key,value in fldcombos.items():
              if fila == sizeoff:
                if num == 1:
                    camposyy += f'{value}'
                elif num == 2:
                    fldselecyy += f'{value}'
                elif num == 3:
                    fldtitulosyy += f'{value}'
                elif num == 4:
                    relacionalyy += f'{value}'
                elif num == 5:
                    camposinsertyy += f'{value}'
                elif num == 6:
                    camposupdateyy += f'{value}'
                elif num == 7:
                    xnomupdateyy += f'{value}'
                elif num == 8:
                    pcamposyy += f'{value}'
                elif num == 9:
                    pcamposoutyy += f'{value}'
                elif num == 10:
                    xcamposinsertyy += f'{value}'
              else:
                if num == 1:
                    camposyy += f'{value}, ' 
                elif num == 2:
                    fldselecyy += f'{value}, '
                elif num == 3:
                    fldtitulosyy += f'{value}, '
                elif num == 4:
                    relacionalyy += f'{value}, '
                elif num == 5:
                    camposinsertyy += f'{value}, '
                elif num == 6:
                    camposupdateyy += f'{value}, '
                elif num == 7:
                    xnomupdateyy += f'{value}, '
                elif num == 8:
                    pcamposyy += f'{value}, '
                elif num == 9:
                    pcamposoutyy += f'{value}, '
                elif num == 10:
                    xcamposinsertyy += f'{value}, '
                    
                fila += 1
        #numtbl,nomtbl,tittbl        
        id = self.request.POST.get('numtbl')   
            
        titulo = self.request.POST.get('tittbl')   
               
        datacombos = f"INSERT INTO tareas_combos(id,nombre,titulo,campos,fldselect,fldtitulos,relacional,camposinsert,camposupdate,xnomupdate,pcampos,pcamposout,xcamposinsert) VALUES ({id},{comi}{nomtbl}{comi},{comi}{titulo}{comi},'','','','','','','','','',''); UPDATE tareas_combos SET campos='{camposyy}', fldselect='{fldselecyy}', fldtitulos='{fldtitulosyy}', relacional='{relacionalyy}', camposinsert='{camposinsertyy}', camposupdate='{camposupdateyy}', xcamposinsert='{xcamposinsertyy}', xnomupdate='{xnomupdateyy}', pcampos='{pcamposyy}', pcamposout='{pcamposoutyy}' WHERE nombre = '{nomtbl}';"
                    
        return datacombos
       except Exception as e:
         print(e) 

class Archivo(Diccionario):
    def __init__(self, request, cmdtext, extfile):
        self.extfile = extfile
        super().__init__(request, cmdtext)
         
    def filaexcel(self, fileexcel, hoja):  
        wb = openpyxl.load_workbook(fileexcel)
        # Define variable to read sheet
        ws = wb[hoja] 
        
        _row = {"titulo": [ws["A1"].value], "descri": [ws["A2"].value]}
        
        if hoja == "Hoja4":
            _row = {"titulo": ws["A1"].value, "descri": [ws["A2"].value,ws["B2"].value,ws["C2"].value,ws["D2"].value,ws["E2"].value,ws["F2"].value,ws["G2"].value]} 
                        
        return _row

    def get_arch(self, titulo): 
            
        userID = self.request.session['user_list'][0]
        
        form = FicheroForm(self.request.POST, self.request.FILES)
        
        context = {'archivo': '', 'is_valid': False}
        
        if form.is_valid():
            is_valid = True
            files = self.request.FILES.getlist('pic')
            
            for arch in files:
                if Path(str(arch)).suffix in self.extfile:
                    arch_ins = Fichero(pic = arch, titulo = titulo, user_id = userID)
                    arch_ins.save() 
            
            dictpdfs = super().get_fila()
            
            pdfs = []
            for ki,vi in dictpdfs.items():
                if ki == 'pic':
                    pdfr = str(vi).replace('_', ' ')
                    pdfs.append(f'media/{pdfr}')
            
            if len(pdfs):
                pdfsort = sorted(pdfs)
            else:   
                return context
                
            if self.request.POST['archi'] == 'nomunir':  
                fusionador = PdfMerger()
                for pdf in pdfsort:
                    pdfr = str(pdf).replace(' ', '_')
                    fusionador.append(pdfr)  
             
                output = f'media/Archivos/{self.request.POST["nomunir"]}.pdf'
                fusionador.write(output)
                fusionador.close() 
                
                context = {'archivo': output, 'is_valid': True}
                      
            elif self.request.POST['archi'] == 'convexcel1':
                 
                pdf_file_path = str(pdfsort[0]).replace(' ', '_') 
                output = f'media/Archivos/{titulo}.xlsx'
                tables = tabula.read_pdf(pdf_file_path, pages='all')
                with pd.ExcelWriter(output) as writer:
                    for i, table in enumerate(tables):
                        table.to_excel(writer, sheet_name=f'Sheet{i+1}')
                        
                headers = ['Fecha','Referencia','Código','Descripción','Débito','CréditoSaldo']
                 
                context = {"archivo": output, 'is_valid': True}
                
                return context 
                      
            elif self.request.POST['archi'] == 'convexcel':
                 
                pdf_file_path = str(pdfsort[0]).replace(' ', '_') 
                output = f'media/Archivos/{titulo}.csv'
                # Create an object of the PdfDocument class
                print(datetime.now())
                tables = tabula.read_pdf(pdf_file_path, pages='all')
                # Set the conversion options
                print(datetime.now())
                tabula.convert_into(pdf_file_path, output, output_format='csv', pages='all')
                print(datetime.now(), 'Convirtiendo PDF a CSV')
                with open(output, newline='') as csvfile:
                    reader = csv.DictReader(csvfile)
                    data = [row for row in reader]
                print(datetime.now(), 'Convirtiendo CSV a JSON')
                outputjson = f'media/Archivos/{titulo}.json'
                with open(outputjson, 'w') as jsonfile:
                    json.dump(data, jsonfile) 
                print(datetime.now(), 'Convirtiendo JSON a datacoopdb')  
                
                with open(outputjson) as fp:
                    data = json.load(fp)
                    
                numregs = super().countreg(98) 
                listerror = []
                for dictdata in data: 
                    dictdataux = super().renameKeysToLower(dictdata)
                    keys = list(dictdataux)
                    cod = dictdataux[keys[0]]                           
                    try:
                        user = User.objects.create_user(cod,email=None,password='1234')
                        user.save()
                        user.first_name = '5'
                        user.save()
                        dictdataux['coop'] = userID
                        dictdataux['user_id'] = user.id
                        insvaldolar = super().fnt_insert(140,dictdataux)   
                    except Exception as e:
                        listerror.append(dictdataux)
                dicterror = {"error": listerror}        
                df = pd.DataFrame(dicterror)
                # Convert dataframe to csv file. The file will pop up on the left in the Files folder
                print(datetime.now(), 'Convirtiendo errores a CSV')
                socioserrorcsv = 'media/Archivos/socioserror.csv'
                df.to_csv(socioserrorcsv, sep=',', index=False, encoding='utf-8')
                print(datetime.now(), 'Convirtiendo errores CSV a JSON')
                dir = f'media/Archivos/'
                socioserror = f"socioserror.json"
                with open(os.path.join(dir, socioserror), 'w') as file:
                    json.dump(listerror, file)
                print(datetime.now(), 'Listo')
                    
                headers = ['Fecha','Referencia','Código','Descripción','Débito','CréditoSaldo']
                 
                context = {"archivo": outputjson, "socioserror": f'{dir}{socioserror}', "socioserrorcsv": socioserrorcsv, 'is_valid': True}
                
                return context 
                 
            elif self.request.POST['archi'] == 'convedacoop': 
                arch = str(pdfsort[0]).replace(' ', '_')
                fileexcel = f"{arch}"
                
                # Read the Excel file                            
                df = pd.read_excel(fileexcel, sheet_name='Hoja1')
                # Convert the DataFrame to JSON
                json_data = df.to_json(orient='records')
                
                # Write the JSON data to a file
                dir = 'media/Archivos/' 
                
                file_name = Path(fileexcel).stem
                file_name = f"{file_name}.json" 
                with open(os.path.join(dir, file_name), 'w', encoding='ISO-8859-1') as file:
                    json.dump(json_data, file) 
                
                # Define variable to load the dataframe
                excel_dataframe = openpyxl.load_workbook(fileexcel)
                
                # Define variable to read sheet
                dataframe = excel_dataframe.active  

                json_data = []
                titulos = []
                # Iterate the loop to read the cell values
                json_data2 = {"banco": [ws["A1"].value]}
                json_data3 = {"lapso": [ws["A2"].value]}
                json_data4 = {"titulo": [ws["B3"].value,ws["C3"].value,ws["D3"].value,ws["E3"].value,ws["F3"].value,ws["G3"].value,ws["H3"].value]} 
                       
                for row in range(3, dataframe.max_row):                   
                    _row = {}
                    j=0
                    for col in dataframe.iter_cols(1, dataframe.max_column):
                       
                        _row[titulos[j]] = valor
                        
                        j += 1
                             
                    json_data.append(_row)
                
                dictdata = [json_data, json_data2, json_data3, json_data4]
                
                dataobj = super().fnt_insert(101, {"archivo": dictdata})
                #print(context)
                print('max_row',dataframe.max_row)    
                context = {"archivo": f'{dir}{file_name}', 'is_valid': True}
                
                return context 
            
            elif self.request.POST['archi'] == 'convedacoop1': 
                  
                # Read the Excel file
                arch = str(pdfsort[0]).replace(' ', '_')
                fileexcel = f"media/Archivos/{arch}"
                json_data2 = self.filaexcel(fileexcel, 'Hoja2')
                json_data3 = self.filaexcel(fileexcel, 'Hoja3')
                json_data4 = self.filaexcel(fileexcel, 'Hoja4')
                
                df = pd.read_excel(fileexcel, sheet_name='Hoja1')
                # Convert the DataFrame to JSON
                json_data = df.to_json(orient='records')
                
                # Write the JSON data to a file
                dir = 'media/Archivos/' 
                file_name = Path(fileexcel).stem
                file_name = f"{file_name}.json"
                with open(os.path.join(dir, file_name), 'w', encoding='ISO-8859-1') as file:
                    json.dump(json_data, file) 
                json_data = json_data.replace('[','')
                json_data = json_data.replace(']','')
                thisdata = {}
                
                #thisdata = {"archivo": json_data, "json_data2": json_data2, "json_data3": json_data3, "json_data4": json_data4}
                
                #dataobj = super().fnt_insert(101, {"archivo": json_data})
                    
                context = {"archivo": f'{dir}{file_name}', 'is_valid': True}
                  
        return context 
		
class Tabla(Diccionario):

    def __init__(self, request, cmdtext):
       super().__init__(request, cmdtext)
        
    def tbl1(self, numtbl):
        
        contexto = super().dict_selectuno()
        
        contexto['titulos'] = contexto['fldtitulos'].split(", ")
        
        style = super().html(1)
        
        tabla = '<table id="tblbase" {style[0]}>'
        thead = '<thead><tr><th scope="col">#</th>'
        
        for titulo in context:
            thead += f'<th scope="col">{titulo}</th>'
            
        thead +=f'</tr></thead>'
        
        context = super().dict_selectodo()        
        
        tbody = '<tbody>'
        for this_dict in context:
            for key,value in this_dict.items():
                tbody += f'<tr>{value}</tr>'
                
        tabla = f'{thead}{tbody}</tbody></table>'
        
        return tabla
        
class Commandtext:
    def __init__(self, ruta, pnumtbl):
        self.ruta = ruta
    
    def xruta():
        cmdtext = 'pcrit_search, pfldbus, pfldord, paccion, ppagant, pindice, plimit, porderby, ptext, puser'    		
class MySuper(object):
    def __init__(self,a,b):
        self.a = a
        self.b = b

class MySub(MySuper):
    def __init__(self,a,b,c):
        self.c = c
        super().__init__(a,b)
'''
my_sub = MySub(42,'chickenman','chuco')
print(my_sub.a)
print(my_sub.b)
print(my_sub.c)
		
testdict = {"archivo": "[{\"Nro.\":1,\"Fecha\":\"02-05-2024\",\"Referencia\":135028952.0,\"C\\u00f3digo\":\"NC\",\"Concepto\":\"PAGO MOVIL P2C O.BANCO VAR.\",\"D\\u00e9bito\":0.0,\"Cr\\u00e9dito\":364.7,\"Saldo\":46801.82},{\"Nro.\":876,\"Fecha\":\"21-05-2024\",\"Referencia\":54635763.0,\"C\\u00f3digo\":\"ND\",\"Concepto\":\"COM PAGO MOVIL P2C O.BANCO VAR\",\"D\\u00e9bito\":2.19,\"Cr\\u00e9dito\":0.0,\"Saldo\":5913.86},{\"Nro.\":877,\"Fecha\":\"21-05-2024\",\"Referencia\":54635763.0,\"C\\u00f3digo\":\"ND\",\"Concepto\":\"P2C COM.GASTOS ADMINIST.\",\"D\\u00e9bito\":0.2,\"Cr\\u00e9dito\":0.0,\"Saldo\":5913.66}]", "json_data2": {"titulo": ["Banco"], "descri": ["Bicentenario Banco Universal"]}, "json_data3": {"titulo": ["Movimientos"], "descri": ["01-05-2024 al 22-05-2024"]}, "json_data4": {"titulo": "T\u00edtulos", "descri": ["Nro.", "Fecha", "Referencia", "C\u00f3digo", "Concepto", "D\u00e9bito", "Cr\u00e9dito", "Saldo"]}, "is_valid": True}
for key,value in testdict.items():
    print(key,value)
   
columns = ['i','double','square']
rows = []

for i in range(6):
    row = [i, i*2, i*i]
    rows.append(row)

df = pd.DataFrame(rows, columns=columns) 
print(df)
   
y = random.sample(range(1, 100), 50)
y1 = [f**2 for f in y]
y2 = [round(math.sqrt(f2)) for f2 in y]
whole = np.stack([y1, y2], axis=-1)
df = pd.DataFrame(whole, index=y, columns=['Square','Square Root'])	
print(df)	

data = np.array([[[1,3],[5,7],[9,11]],[[2,4],[6,8],[10,12]]])
df = pd.DataFrame(data.tolist(), columns=['A','B','C'])	
print(df)			

a = pd.Series([['a','c'],[1,3]])
b = pd.Series([['b','d'],[2,4]])

df = pd.DataFrame([a,b])

print(df)		
'''





